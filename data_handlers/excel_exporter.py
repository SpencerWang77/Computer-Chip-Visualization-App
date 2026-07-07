import os
from datetime import datetime

import openpyxl


class ExcelExporter:
    """Writes the (possibly modified) pad data to a new Excel file.

    The original workbook is never modified; the caller chooses the output
    path. Multi-die exports go into one combined sheet whose coordinates are
    in the shared ring coordinate system (origin at the ring's bottom-left).
    """

    # Mirror the input die-netlist tabs exactly (row 1 headers, row 2 blank),
    # so exported die tabs are structurally identical to the input.
    HEADERS = ["Die Pad No", "Pad name", "X-coord\n(after shrink)",
               "Y-coord\n(after shrink)", "X open\n(after shrink)",
               "Y open\n(after shrink)", "Net Name", "Bonding\nrelationship"]
    UNITS = ["", "", "", "", "", "", "", ""]

    def __init__(self, source_file=None):
        self.source_file = source_file

    def export_by_die(self, rows, output_path, metadata=None):
        """Export one 'Die Netlist(<name>)' tab per die (mirroring the input),
        with pad coordinates in each die's own local frame, plus a 'Basic
        information' sheet recording the VSS ring size and every die's placement
        so the file can be reopened exactly where it was left off.

        `rows` is a list of (die_name, Pad). Returns (success, saved_path).
        """
        try:
            if not output_path:
                raise ValueError("No output path given")
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)

            by_die = {}
            for die_name, pad in rows:
                by_die.setdefault(die_name, []).append(pad)

            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # drop the default empty sheet

            if metadata:
                self._write_basic_information(workbook, metadata)

            for die_name, pads in by_die.items():
                sheet = workbook.create_sheet(title=self._sheet_title(die_name))
                for col, value in enumerate(self.HEADERS, 1):
                    sheet.cell(row=1, column=col, value=value)
                for col, value in enumerate(self.UNITS, 1):
                    sheet.cell(row=2, column=col, value=value)
                for row_idx, pad in enumerate(pads, 3):
                    sheet.cell(row=row_idx, column=1, value=pad.pad_id)
                    sheet.cell(row=row_idx, column=2, value=pad.pad_name)
                    sheet.cell(row=row_idx, column=3, value=pad.x_coord)
                    sheet.cell(row=row_idx, column=4, value=pad.y_coord)
                    sheet.cell(row=row_idx, column=5, value=pad.x_open)
                    sheet.cell(row=row_idx, column=6, value=pad.y_open)
                    sheet.cell(row=row_idx, column=7, value=pad.net_name)
                    sheet.cell(row=row_idx, column=8, value=pad.bonding)

            workbook.save(output_path)
            return True, output_path

        except Exception as e:
            print(f"Excel export failed: {e}")
            return False, None

    def _write_basic_information(self, workbook, meta):
        """Machine-readable layout state, read back by ExcelHandler."""
        sheet = workbook.create_sheet(title="Basic information")
        sheet.cell(row=1, column=1, value="Basic information")
        sheet.cell(row=1, column=2, value="(saved by Chip Pad Editor — reopen to resume)")

        sheet.cell(row=3, column=1, value="VSS ring size (um)")
        sheet.cell(row=3, column=2, value=meta.get('ring_w'))
        sheet.cell(row=3, column=3, value=meta.get('ring_h'))

        sheet.cell(row=4, column=1, value="Lead frame pins")
        sheet.cell(row=4, column=2, value=meta.get('pin_x'))
        sheet.cell(row=4, column=3, value=meta.get('pin_y'))

        sheet.cell(row=6, column=1, value="Die placement (ring coordinates)")
        header = ["Die", "Center X", "Center Y", "Rotation (deg)", "Width", "Height"]
        for col, value in enumerate(header, 1):
            sheet.cell(row=7, column=col, value=value)
        for i, die in enumerate(meta.get('dies', []), 8):
            sheet.cell(row=i, column=1, value=die['name'])
            sheet.cell(row=i, column=2, value=round(die['center_x'], 4))
            sheet.cell(row=i, column=3, value=round(die['center_y'], 4))
            sheet.cell(row=i, column=4, value=round(die['angle'], 4))
            sheet.cell(row=i, column=5, value=round(die['width'], 4))
            sheet.cell(row=i, column=6, value=round(die['height'], 4))

    @staticmethod
    def _sheet_title(die_name):
        # Excel sheet titles are capped at 31 chars.
        return f"Die Netlist({die_name})"[:31]

    def get_export_summary(self, pads):
        return {
            'total_pads': len(pads),
            'modified_pads': len([p for p in pads if p.is_modified]),
            'export_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
