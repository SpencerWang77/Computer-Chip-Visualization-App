import re
from typing import List

import openpyxl

from models import Die, Pad

# Matches bonding values like "LF.119", "LF 12", "LF_7"
_LF_PATTERN = re.compile(r'^LF[._\s]*(\d+)$', re.IGNORECASE)
# Matches a die netlist tab name like "Die Netlist(SOC)"
_DIE_TAB_PATTERN = re.compile(r'^\s*Die\s*Netlist\s*\((.+)\)\s*$', re.IGNORECASE)
# Two numbers in a size string like "3475.8*3105um(...)"
_SIZE_PATTERN = re.compile(r'([\d.]+)\s*[*x×]\s*([\d.]+)')


class ExcelHandler:
    """Reads chip data from an Excel workbook.

    Two layouts are supported:
      - Multi-die: one sheet per die named "Die Netlist(<name>)", plus an
        optional "Basic information" sheet giving die sizes and pin count.
      - Legacy single-die: a sheet named "connect".

    Each die sheet has: row 1 headers, row 2 units/blank, row 3+ pad data
    (Die Pad No, Pad name, X-coord, Y-coord, X open, Y open, Net Name, Bonding).
    """

    def __init__(self):
        self.dies: List[Die] = []
        self.source_file = None
        self.pin_count = None
        self.saved_pins = None       # (pin_x, pin_y) from a saved layout
        self.saved_ring_size = None  # (width, height) from a saved layout
        self.saved_placements = {}   # die_name -> (center_x, center_y, angle)

    def read_excel(self, file_path: str) -> bool:
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            die_tabs = [(name, self._DIE_name(name)) for name in workbook.sheetnames
                        if _DIE_TAB_PATTERN.match(name)]

            sizes = self._read_basic_information(workbook)

            self.dies = []
            if die_tabs:
                for sheet_name, die_name in die_tabs:
                    pads = self._read_pads(workbook[sheet_name])
                    if not pads:
                        continue
                    w, h = sizes.get(die_name, (None, None))
                    self.dies.append(Die(die_name, pads, w, h))
            elif 'connect' in workbook.sheetnames:
                pads = self._read_pads(workbook['connect'])
                if pads:
                    die_name = pads[0].pad_id.split('.')[0] or "DIE"
                    self.dies.append(Die(die_name, pads))
            else:
                raise ValueError("No 'Die Netlist(<name>)' or 'connect' sheet found")

            if not self.dies:
                raise ValueError("No pad data found in any die sheet")

            self.source_file = file_path
            return True

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return False

    @staticmethod
    def _DIE_name(sheet_name: str) -> str:
        return _DIE_TAB_PATTERN.match(sheet_name).group(1).strip()

    @staticmethod
    def _is_number(v) -> bool:
        try:
            float(v)
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _read_pads(sheet) -> List[Pad]:
        """Read pad rows, auto-detecting where the data starts. A data row has
        a non-empty first cell (the pad id) and a numeric X-coord, so any
        number of header/title/units rows is skipped — this keeps the first
        pad (e.g. SOC.1) from being dropped when the title layout varies."""
        pads = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            pad_id = row[0].value if len(row) > 0 else None
            x_coord = row[2].value if len(row) > 2 else None
            if not pad_id or not ExcelHandler._is_number(x_coord):
                continue  # header / units / title / blank row
            pad_name = row[1].value if len(row) > 1 else ""
            y_coord = row[3].value if len(row) > 3 else 0
            x_open = row[4].value if len(row) > 4 else 0
            y_open = row[5].value if len(row) > 5 else 0
            net_name = row[6].value if len(row) > 6 else ""
            bonding = row[7].value if len(row) > 7 else ""
            pads.append(Pad(pad_id, pad_name, x_coord, y_coord,
                            x_open, y_open, net_name, bonding))
        return pads

    def _read_basic_information(self, workbook):
        """Parse the 'Basic information' sheet, supporting both the original
        input format (DIE Code / Die size / Pin count) and the format this app
        writes on export (VSS ring size / Lead frame pins / Die placement).

        Populates self.pin_count, self.saved_pins, self.saved_ring_size and
        self.saved_placements; returns {die_name: (width, height)}."""
        sizes = {}
        self.pin_count = None
        self.saved_pins = None
        self.saved_ring_size = None
        self.saved_placements = {}
        if 'Basic information' not in workbook.sheetnames:
            return sizes
        sheet = workbook['Basic information']
        rows = list(sheet.iter_rows(values_only=True))

        def num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        die_codes = []  # column index (>=1) -> die name (original format)
        size_row = None
        for r, row in enumerate(rows):
            label = str(row[0]).strip().lower() if row and row[0] else ""

            # --- original input format ---
            if 'die code' in label:
                die_codes = [(i, str(v).strip()) for i, v in enumerate(row)
                             if i >= 1 and v]
            elif 'die size' in label:
                size_row = row
            elif 'pin count' in label:
                for v in row[1:]:
                    if v and re.search(r'\d', str(v)):
                        self.pin_count = int(re.search(r'\d+', str(v)).group())
                        break

            # --- this app's saved layout ---
            elif 'vss ring size' in label:
                w, h = num(row[1] if len(row) > 1 else None), num(row[2] if len(row) > 2 else None)
                if w and h:
                    self.saved_ring_size = (w, h)
            elif 'lead frame pins' in label:
                x, y = num(row[1] if len(row) > 1 else None), num(row[2] if len(row) > 2 else None)
                if x and y:
                    self.saved_pins = (int(x), int(y))
            elif 'die placement' in label:
                self.saved_placements = self._read_placement_table(rows, r + 1, sizes)

        if die_codes and size_row is not None:
            for col, name in die_codes:
                if col < len(size_row) and size_row[col]:
                    m = _SIZE_PATTERN.search(str(size_row[col]))
                    if m:
                        sizes[name] = (float(m.group(1)), float(m.group(2)))
        return sizes

    @staticmethod
    def _read_placement_table(rows, start, sizes):
        """Parse the 'Die placement' table: a header row then
        (name, cx, cy, angle, width, height) rows until a blank name."""
        placements = {}
        for row in rows[start:]:
            if not row or not row[0]:
                break
            first = str(row[0]).strip().lower()
            if first in ('die', 'name'):  # header row
                continue

            def cell(i):
                try:
                    return float(row[i]) if len(row) > i and row[i] is not None else None
                except (TypeError, ValueError):
                    return None

            name = str(row[0]).strip()
            cx, cy, angle = cell(1), cell(2), cell(3)
            w, h = cell(4), cell(5)
            if cx is not None and cy is not None:
                placements[name] = (cx, cy, angle or 0.0)
            if w and h:
                sizes[name] = (w, h)
        return placements

    # --- accessors -------------------------------------------------------

    def get_dies(self) -> List[Die]:
        return self.dies

    def get_saved_ring_size(self):
        """(width, height) if the file was saved by this app, else None."""
        return self.saved_ring_size

    def get_saved_pins(self):
        """(pin_x, pin_y) if the file was saved by this app, else None."""
        return self.saved_pins

    def get_saved_placements(self):
        """{die_name: (center_x, center_y, angle)} from a saved layout, or {}."""
        return dict(self.saved_placements)

    def get_pads(self) -> List[Pad]:
        """All pads across every die (flattened)."""
        return [pad for die in self.dies for pad in die.pads]

    def get_max_lf_pin(self) -> int:
        max_pin = 0
        for pad in self.get_pads():
            match = _LF_PATTERN.match(pad.bonding)
            if match:
                max_pin = max(max_pin, int(match.group(1)))
        return max_pin

    def suggest_pin_count(self) -> int:
        """Pin count from 'Basic information', else the highest LF.<n> rounded
        up to a multiple of 4."""
        if self.pin_count:
            return self.pin_count
        max_pin = self.get_max_lf_pin()
        return ((max_pin + 3) // 4) * 4 if max_pin else 0
