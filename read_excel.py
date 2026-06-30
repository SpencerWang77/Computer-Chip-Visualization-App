import openpyxl

wb = openpyxl.load_workbook('封装连线示意.xlsx')
print("工作表:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n工作表: {sheet_name}")
    print(f"数据范围: {ws.max_row} 行 x {ws.max_column} 列")
    
    if ws.max_row > 0:
        print("前10行数据:")
        for i in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for j in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=i, column=j).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"  第{i}行: {row_data}")
    else:
        print("工作表为空")
